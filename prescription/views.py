from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render, redirect
from .forms import CustomUserCreationForm,PrescriptionForm,MedicineUsageForm,PatientForm,InitialExaminationForm,ProfileForm,FileUploadForm
from django.contrib.auth.models import User
from .models import Prescription,Patient,Medicine,MedicineTypes,MedicineUsage,Advice,Investigation,InitialExamination,FileUpload
from openpyxl import load_workbook
# Create your views here.

@login_required(login_url='login')
def profile(request):
    user_profile = request.user.profile
    context={
        'profile':user_profile
    }
    return render(request,'prescription/profile.html',context)


@login_required(login_url='login')
def profile_edit(request):
    profile = request.user.profile
    profile_form = ProfileForm(instance=profile)
    if request.method == "POST":
        profile_form = ProfileForm(request.POST,instance=profile)
        if profile_form.is_valid():
            profile_form.save()
            return redirect('profile')
    context={
        'form':profile_form
    }
    return render(request,'prescription/profile-edit.html',context)

def medicine_bulk_add(request):
    form = FileUploadForm()
    if request.method == "POST":
        form = FileUploadForm(request.POST,request.FILES)
        if form.is_valid():
            form.save()
            excel = FileUpload.objects.get(used=False)
            wb = load_workbook(excel.file.path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=sheet.min_row+1,
                                       max_row=sheet.max_row,
                                       min_col=sheet.min_column,
                                       max_col=sheet.max_column):
                val = [data.value for data in row]
                name = val[0]
                group = val[1]
                type = val[2]

                if type == MedicineTypes.Tablet or type == MedicineTypes.Injection or type == MedicineTypes.Saline:
                    type = MedicineTypes.Tablet
                elif type == MedicineTypes.Syrup:
                    type = MedicineTypes.Syrup
                elif type == MedicineTypes.Injection:
                    type = MedicineTypes.Injection
                elif type == MedicineTypes.Saline:
                    type = MedicineTypes.Saline
                else:
                    type = MedicineTypes.Tablet

                Medicine.objects.get_or_create(
                    name=name,
                    group=group,
                    type=type
                )
            excel.used = True
            excel.save()

    context={
        'form':form
    }

    return render(request,'prescription/bulk-add.html',context)


def advice_bulk_add(request):
    form = FileUploadForm()
    if request.method == "POST":
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            excel = FileUpload.objects.get(used=False)
            wb = load_workbook(excel.file.path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=sheet.min_row + 1,
                                       max_row=sheet.max_row,
                                       min_col=1,
                                       max_col=1):
                for data in row:
                    val = data.value
                    Advice.objects.get_or_create(doctor=request.user.profile,title=val)

            excel.used = True
            excel.save()

    context = {
        'form': form
    }

    return render(request, 'prescription/bulk-add.html', context)

@login_required(login_url='login')
def prescription_create(request):
    prescription = Prescription.objects.create()
    patient = Patient.objects.create()
    initial_examination = InitialExamination.objects.create()
    prescription.patient = patient
    prescription.initial_examination = initial_examination
    prescription.save()
    return redirect('prescription-update',prescription.id)


@login_required(login_url='login')
def prescription_update(request,pk):
    print(request.POST)
    prescription_instance = Prescription.objects.get(id=pk)
    patient_instance = Patient.objects.get(id=prescription_instance.patient.id)
    initial_examination_instance = InitialExamination.objects.get(id=prescription_instance.initial_examination.id)

    prescription_form = PrescriptionForm(instance=prescription_instance)
    patient_form = PatientForm(instance=patient_instance)
    initial_examination_form = InitialExaminationForm(instance=initial_examination_instance)
    medicine_usage_form = MedicineUsageForm()

    if request.method == 'POST':
        advices = request.POST.get("generalAdvice").replace(', ', ',').split(",")
        advices.pop()

        investigations = request.POST.get("investigation").replace(', ', ',').split(",")
        investigations.pop()

        patient_data={
            'name':request.POST['name'],
            'gender':request.POST['gender'],
            'age':request.POST['age'],
            'phone_number':request.POST['phone_number']
        }

        prescription_data = {
            'complaints': request.POST['complaints'],
            'diagnosis': request.POST['diagnosis'],
        }

        initial_examination_data = {
            'complaints': request.POST['complaints'],
            'diagnosis': request.POST['diagnosis'],
        }

        prescription_form = PrescriptionForm(prescription_data,instance=prescription_instance)
        patient_form = PatientForm(patient_data,instance=patient_instance)

        initial_examination_form = InitialExaminationForm(request.POST,instance=initial_examination_instance)

        if prescription_form.is_valid() and patient_form.is_valid() and initial_examination_form.is_valid():

            prescription = prescription_form.save(commit=False)
            patient = patient_form.save()
            initial_examination = initial_examination_form.save()
            prescription.doctor = request.user.profile
            prescription.patient = patient
            prescription.initial_examination = initial_examination
            prescription.save()

            for advice in advices:
                advice, created = Advice.objects.get_or_create(title=advice, doctor=request.user.profile)
                prescription.advice.add(advice)

            for investigation in investigations:
                investigation, created = Investigation.objects.get_or_create(name=investigation)
                prescription.investigation.add(investigation)

            # medicines = MedicineUsage.objects.filter(prescription__isnull=True)
            # for medicine in medicines:
            #     medicine.prescription = prescription
            #     medicine.save()

            # prescription_form = PrescriptionForm()
            # medicine_usage_form = MedicineUsageForm()
            # patient_form = PatientForm()
            # initial_examination_form = InitialExaminationForm()

    context = {
        'prescription_form': prescription_form,
        'medicine_usage_form': medicine_usage_form,
        'patient_form': patient_form,
        'initial_examination_form': initial_examination_form,
        'prescription':prescription_instance,
    }

    return render(request, 'prescription/prescription-create.html', context)


def home(request):
    return render(request,'prescription/home.html')


def register_user(request):
    form = CustomUserCreationForm()

    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = user.username.lower()
            try:
                User.objects.get(username=user.username)
                form = CustomUserCreationForm(request.POST)
                context = {
                    'form': form
                }
                return render(request, 'prescription/registration.html', context)
            except ObjectDoesNotExist:
                user.save()
                login(request, user)
                return redirect('profile-edit')
        else:
            form = CustomUserCreationForm(request.POST)
            context = {
                'form': form
            }
            return render(request, 'prescription/registration.html', context)

    context = {
        'form': form
    }

    return render(request,'prescription/registration.html',context)


def login_user(request):

    if request.user.is_authenticated:
        return redirect('home')

    if request.method == "POST":
        username = request.POST['username'].lower()
        password = request.POST['password']

        try:
            user = User.objects.get(username=username)
        except ObjectDoesNotExist:
            return redirect('login')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("home")
        else:
            return redirect('login')

    return render(request, 'prescription/login.html')


@login_required(login_url='login')
def logout_user(request):
    logout(request)
    return redirect('login')